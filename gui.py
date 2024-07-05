import tkinter as tk
from tkinter import messagebox, filedialog
import subprocess
import json
import os
from fpdf import FPDF
import pandas as pd
from ctypes import windll
import sys
import time
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


class SystemAnalyzerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("System Analyzer")

        # Title
        self.title_label = tk.Label(
            root, text="Scanning Options", font=("Helvetica", 20)
        )
        self.title_label.pack(pady=10)

        self.frame = tk.Frame(root)
        self.frame.pack(pady=20)

        # Analysis options
        self.options = {
            "System Information": tk.BooleanVar(),
            "Windows Defender Version": tk.BooleanVar(),
            "Specific firewall rules are enforced": tk.BooleanVar(),
            "System Update Status": tk.BooleanVar(),
            "BitLocker Status": tk.BooleanVar(),
            "Open TCP Ports": tk.BooleanVar(),
            "USB Malware Protection": tk.BooleanVar(),
            "Pending Updates": tk.BooleanVar(),
            "DNS Cache Check": tk.BooleanVar(),
            "System Logs": tk.BooleanVar(),
            "User Account Audit": tk.BooleanVar(),
            "File Integrity Check": tk.BooleanVar(),
            "VPN and Proxy Check": tk.BooleanVar(),
            "Anonymous Connections": tk.BooleanVar(),
        }

        # Create checkboxes and place them in two columns
        row = 0
        col = 0
        for option, var in self.options.items():
            checkbutton = tk.Checkbutton(
                self.frame, text=option, variable=var, font=("Helvetica", 13)
            )
            checkbutton.grid(
                row=row,
                column=col,
                sticky="w",
                padx=10,
                pady=10,
            )
            row += 1
            if row >= len(self.options) // 2:
                row = 0
                col += 1

        self.run_button = tk.Button(
            self.frame,
            text="Run Analysis",
            command=self.run_analysis,
            font=("Helvetica", 16),
        )
        self.run_button.grid(row=len(self.options) // 2 + 1, column=0, pady=5)

        self.export_button = tk.Button(
            self.frame,
            text="Export Results",
            command=self.export_results,
            font=("Helvetica", 16),
        )
        self.export_button.grid(row=len(self.options) // 2 + 1, column=1, pady=5)

        self.results = None

    def run_analysis(self):
        try:
            # Mapping dos nomes do Tkinter para os nomes dos parâmetros PowerShell
            param_map = {
                "System Information": "-GetSystemInfo",
                "Windows Defender Version": "-CheckWindowsDefender",
                "Specific firewall rules are enforced": "-CheckFirewallRules",
                "System Update Status": "-CheckSystemUpdates",
                "BitLocker Status": "-CheckBitLocker",
                "Open TCP Ports": "-CheckOpenPorts",
                "USB Malware Protection": "-CheckUSBProtection",
                "Pending Updates": "-CheckPendingUpdates",
                "DNS Cache Check": "-CheckDNSCache",
                "System Logs": "-MonitorLogs",
                "User Account Audit": "-AuditUserAccounts",
                "File Integrity Check": "-CheckFileIntegrity",
                "VPN and Proxy Check": "-CheckVPNProxies",
                "Anonymous Connections": "-MonitorAnonConnections",
            }

            # Build the PowerShell command with parameters based on selected options
            ps_command = ["powershell.exe", os.path.abspath("./main.ps1")]
            for option, var in self.options.items():
                if var.get():
                    ps_command.append(param_map[option])
                    print(param_map[option])  # Verifica os parâmetros sendo adicionados

            print("Running command:", ps_command)  # Print para depuração

            result = subprocess.run(
                ps_command, capture_output=True, text=True, check=True
            )

            if result.returncode == 0:
                # Verificar se o arquivo security_report.json foi criado
                max_attempts = 10
                attempts = 0
                file_path = "security_report.json"
                while not os.path.exists(file_path) or os.stat(file_path).st_size == 0:
                    time.sleep(1)  # Espera 1 segundo antes da próxima verificação
                    attempts += 1
                    if attempts >= max_attempts:
                        messagebox.showerror(
                            "Error",
                            "Failed to find valid security_report.json after waiting.",
                        )
                        return

                # Carregar o arquivo JSON após ser criado
                try:
                    with open(file_path, "r", encoding="utf-8-sig") as file:
                        json_data = file.read()
                        if json_data.strip():  # Verifica se há conteúdo no arquivo
                            self.results = json.loads(json_data)
                            messagebox.showinfo(
                                "Analysis Complete",
                                "System analysis completed successfully.",
                            )
                        else:
                            messagebox.showerror(
                                "Error", "security_report.json is empty."
                            )
                except json.JSONDecodeError as e:
                    messagebox.showerror("Error", f"Failed to parse JSON: {str(e)}")
            else:
                messagebox.showerror("Error", "An error occurred during analysis.")
        except subprocess.CalledProcessError as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

    def export_results(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("Excel files", "*.xlsx")],
        )
        if file_path:
            if file_path.endswith(".pdf"):
                self.export_to_pdf(file_path)
            elif file_path.endswith(".xlsx"):
                self.export_to_excel(file_path)

    def load_results_from_json(self, file_path):
        try:
            if os.path.exists(file_path):
                with open(file_path, "r", encoding="utf-8") as f:
                    self.results = json.load(f)
            else:
                messagebox.showerror(
                    "File Not Found", f"The JSON file '{file_path}' was not found."
                )
        except json.JSONDecodeError as e:
            messagebox.showerror(
                "JSON Error", f"Failed to parse JSON file '{file_path}': {e}"
            )
        except FileNotFoundError:
            messagebox.showerror(
                "File Not Found", f"The JSON file '{file_path}' was not found."
            )
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    def export_to_pdf(self, file_path):
        try:
            if not self.results:
                messagebox.showerror(
                    "Export Error", "No results loaded. Please run the analysis first."
                )
                return

            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Helvetica", size=12)

            for key, value in self.results.items():
                if key == "System Information":
                    pdf.set_font("Helvetica", "B", 12)
                    pdf.cell(200, 10, txt=f"{key}:", ln=True)
                    pdf.set_font("Helvetica", size=12)

                    if isinstance(value, dict):
                        for k, v in value.items():
                            pdf.cell(200, 10, txt=f"{k}: {v}", ln=True)
                    else:
                        pdf.cell(200, 10, txt=str(value), ln=True)
                    pdf.ln()

                elif key == "Open TCP Ports":
                    pdf.set_font("Helvetica", "B", 12)
                    pdf.cell(200, 10, text="Open TCP Ports")
                    pdf.ln()
                    headers = [
                        "Local Address",
                        "Remote Port",
                        "Local Port",
                        "Status",
                    ]
                    for header in headers:
                        pdf.cell(40, 10, txt=header, border=1)
                    pdf.ln()
                    pdf.set_font("Helvetica", size=12)
                    if isinstance(value, list):
                        self.add_open_tcp_ports_section(pdf, value)
                    else:
                        pdf.cell(200, 10, txt=f"{key}: {value}")
                        pdf.ln()

                elif key == "Anonymous Connections":
                    pdf.set_font("Helvetica", "B", 12)
                    pdf.cell(200, 10, txt="Anonymous Connections")
                    pdf.ln()
                    headers = [
                        "Local Address",
                        "Remote Address",
                        "Local Port",
                        "Remote Port",
                    ]
                    for header in headers:
                        pdf.cell(40, 10, txt=header, border=1)
                    pdf.ln()
                    pdf.set_font("Helvetica", size=12)
                    if isinstance(value, list):
                        self.add_anonymous_connections_section(pdf, value)
                    else:
                        pdf.cell(200, 10, txt=f"{key}: {value}")
                        pdf.ln()

                elif key == "Specific firewall rules are enforced":
                    pdf.set_font("Helvetica", "B", 12)
                    pdf.cell(200, 10, txt="Firewall Enforced Rules")
                    pdf.ln()

                    headers = ["Profile", "Description", "Name", "Direction", "Action"]
                    for header in headers:
                        pdf.cell(38, 10, txt=header, border=1)
                    pdf.ln()
                    pdf.set_font("Helvetica", size=12)

                    if isinstance(value, list):
                        for rule in value:
                            if isinstance(rule, dict):
                                pdf.cell(38, 10, txt=rule.get("Profile", ""), border=1)
                                pdf.cell(
                                    38, 10, txt=rule.get("Description", ""), border=1
                                )
                                pdf.cell(38, 10, txt=rule.get("Name", ""), border=1)
                                pdf.cell(
                                    38, 10, txt=rule.get("Direction", ""), border=1
                                )
                                pdf.cell(38, 10, txt=rule.get("Action", ""), border=1)
                                pdf.ln()
                            else:
                                pdf.cell(200, 10, txt=str(rule))
                                pdf.ln()

                elif key == "User Account Audit":
                    pdf.set_font("Helvetica", "B", 12)
                    pdf.cell(200, 10, text="User Account Audit")

                    pdf.ln()
                    # Define headers for the table
                    headers = ["Full Name", "Name", "Last Login"]
                    for header in headers:
                        pdf.cell(60, 10, text=header, border=1)
                    pdf.ln()
                    pdf.set_font("Helvetica", size=12)
                    if isinstance(value, dict):
                        pdf.cell(60, 10, txt=value.get("FullName", ""), border=1)
                        pdf.cell(60, 10, txt=value.get("Name", ""), border=1)
                        date = self.converter_data(str(value.get("LastLogon")))
                        pdf.cell(60, 10, txt=date, border=1)
                        pdf.ln()
                    else:
                        pdf.cell(200, 10, txt=f"{key}: {value}")
                        pdf.ln()

                elif key == "DNS Cache Check":
                    pdf.set_font("Helvetica", "B", 12)
                    pdf.cell(200, 10, txt="DNS Cache Check")
                    pdf.set_font("Helvetica", size=12)
                    pdf.ln()

                    if isinstance(value, list) and value:
                        for dns_entry in value:
                            pdf.cell(200, 10, txt=dns_entry, ln=True)
                    else:
                        pdf.cell(200, 10, txt=str(value))
                    pdf.ln()

                elif key == "USB Malware Protection":
                    pdf.set_font("Helvetica", "B", 12)
                    pdf.cell(200, 10, txt="USB Malware Protection")
                    pdf.set_font("Helvetica", size=12)
                    pdf.ln()

                    if isinstance(value, dict):
                        for k, v in value.items():
                            pdf.cell(200, 10, txt=f"{k}: {v}", ln=True)

                elif key == "VPN and Proxy Check":
                    pdf.set_font("Helvetica", "B", 12)
                    pdf.cell(200, 10, txt="VPN and Proxy Connections")
                    pdf.ln()
                    headers = [
                        "Status",
                        "Name",
                        "Address",
                    ]

                    for header in headers:
                        pdf.cell(60, 10, txt=header, border=1)
                    pdf.set_font("Helvetica", size=12)
                    pdf.ln()

                    if isinstance(value, dict):
                        pdf.cell(
                            60, 10, txt=value.get("ConnectionStatus", ""), border=1
                        )
                        pdf.cell(60, 10, txt=value.get("Name", ""), border=1)
                        pdf.cell(60, 10, txt=value.get("ServerAddress", ""), border=1)
                        pdf.ln()

                elif isinstance(value, dict):
                    for k, v in value.items():
                        if isinstance(v, list):
                            for item in v:
                                if isinstance(item, dict):
                                    for ki, vi in item.items():
                                        pdf.cell(200, 10, txt=f"{ki}: {vi}")
                                        pdf.ln()
                                else:
                                    pdf.cell(200, 10, txt=f"{k}: {v}")
                                    pdf.ln()
                        else:
                            pdf.cell(200, 10, txt=f"{k}: {v}")
                            pdf.ln()

                elif isinstance(value, list):
                    for item in value:
                        if isinstance(item, dict):
                            for k, v in item.items():
                                pdf.cell(200, 10, txt=f"{k}: {v}")
                                pdf.ln()
                        else:
                            pdf.cell(200, 10, txt=str(item))
                            pdf.ln()

                elif isinstance(value, str):
                    pdf.cell(200, 10, txt=f"{key}: {value}")
                    pdf.ln()

                else:
                    pdf.cell(200, 10, txt=f"{key}: {value}")
                    pdf.ln()

            pdf.output(file_path)
            messagebox.showinfo("Export Complete", f"Results exported to {file_path}")

        except Exception as e:
            messagebox.showerror(
                "Export Error", f"An error occurred during export: {str(e)}"
            )

    def add_open_tcp_ports_section(self, pdf, data):
        pdf.set_font("Helvetica", size=12)
        for port in data:
            if isinstance(port, dict):
                pdf.cell(40, 10, text=port.get("Local Address", ""), border=1)
                pdf.cell(40, 10, text=str(port.get("Local Port", "")), border=1)
                pdf.cell(40, 10, text=str(port.get("Remote Port", "")), border=1)
                pdf.cell(40, 10, text=port.get("State", ""), border=1)
                pdf.ln()
            else:
                pdf.cell(200, 10, text=str(port))
        pdf.ln()

    def add_anonymous_connections_section(self, pdf, data):
        for connection in data:
            if isinstance(connection, dict):
                pdf.cell(40, 10, text=connection.get("LocalAddress", ""), border=1)
                pdf.cell(40, 10, text=connection.get("RemoteAddress", ""), border=1)
                pdf.cell(40, 10, text=str(connection.get("LocalPort", "")), border=1)
                pdf.cell(40, 10, text=str(connection.get("RemotePort", "")), border=1)
                pdf.ln()
            else:
                pdf.cell(200, 10, text=str(connection))
        pdf.ln()

    def converter_data(self, data_milissegundos):
        # Extrair o número de milissegundos da string
        milissegundos = int(data_milissegundos.strip("\/Date()\/"))

        # Converter milissegundos para segundos
        segundos = milissegundos / 1000

        # Criar um objeto datetime a partir dos segundos
        data_legivel = datetime.datetime.utcfromtimestamp(segundos)

        # Formatar a data para um formato legível
        data_formatada = data_legivel.strftime("%Y-%m-%d %H:%M:%S")

        return data_formatada

    def export_to_excel(self, file_path):
        try:
            if not self.results:
                messagebox.showerror(
                    "Export Error", "No results loaded. Please run the analysis first."
                )
                return

            # Create a new Workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Analysis Results"

            # Set default font and alignment
            default_font = Font(name="Helvetica", size=12)
            default_alignment = Alignment(horizontal="left", vertical="top")

            row = 1  # Start writing from the first row

            for key, value in self.results.items():
                if key == "System Information":
                    ws.merge_cells(
                        start_row=row, start_column=1, end_row=row, end_column=2
                    )
                    ws.cell(row=row, column=1).value = f"{key}: {value}"
                    ws.cell(row=row, column=1).font = Font(bold=True)
                    row += 1

                elif key == "Open TCP Ports":
                    start_merge_row = row
                    headers = ["Local Address", "Remote Port", "Local Port", "Status"]
                    for idx, header in enumerate(headers, start=1):
                        ws.cell(row=row, column=idx).value = header
                        ws.cell(row=row, column=idx).font = Font(bold=True)
                    row += 1
                    if isinstance(value, list):
                        row = self.add_open_tcp_ports_to_excel(ws, value, row)
                    else:
                        ws.cell(row=row, column=1).value = value
                        ws.cell(row=row, column=1).font = default_font
                        row += 1
                    end_merge_row = row - 1
                    ws.merge_cells(
                        start_row=start_merge_row,
                        start_column=1,
                        end_row=end_merge_row,
                        end_column=4,
                    )

                elif key == "Anonymous Connections":
                    start_merge_row = row
                    headers = [
                        "Local Address",
                        "Remote Address",
                        "Local Port",
                        "Remote Port",
                    ]
                    for idx, header in enumerate(headers, start=1):
                        ws.cell(row=row, column=idx).value = header
                        ws.cell(row=row, column=idx).font = Font(bold=True)
                    row += 1
                    if isinstance(value, list):
                        row = self.add_anonymous_connections_to_excel(ws, value, row)
                    else:
                        ws.cell(row=row, column=1).value = value
                        ws.cell(row=row, column=1).font = default_font
                        row += 1
                    end_merge_row = row - 1
                    ws.merge_cells(
                        start_row=start_merge_row,
                        start_column=1,
                        end_row=end_merge_row,
                        end_column=4,
                    )

                elif key == "Specific firewall rules are enforced":
                    start_merge_row = row
                    headers = ["Profile", "Description", "Name", "Direction", "Action"]
                    for idx, header in enumerate(headers, start=1):
                        ws.cell(row=row, column=idx).value = header
                        ws.cell(row=row, column=idx).font = Font(bold=True)
                    row += 1
                    if isinstance(value, list):
                        row = self.add_firewall_rules_to_excel(ws, value, row)
                    else:
                        ws.cell(row=row, column=1).value = value
                        ws.cell(row=row, column=1).font = default_font
                        row += 1
                    end_merge_row = row - 1
                    ws.merge_cells(
                        start_row=start_merge_row,
                        start_column=1,
                        end_row=end_merge_row,
                        end_column=5,
                    )

                elif key == "User Account Audit":
                    start_merge_row = row
                    headers = ["Full Name", "Name", "Last Login"]
                    for idx, header in enumerate(headers, start=1):
                        ws.cell(row=row, column=idx).value = header
                        ws.cell(row=row, column=idx).font = Font(bold=True)
                    row += 1
                    if isinstance(value, dict):
                        row = self.add_user_account_audit_to_excel(ws, value, row)
                    else:
                        ws.cell(row=row, column=1).value = value
                        ws.cell(row=row, column=1).font = default_font
                        row += 1
                    end_merge_row = row - 1
                    ws.merge_cells(
                        start_row=start_merge_row,
                        start_column=1,
                        end_row=end_merge_row,
                        end_column=3,
                    )

                elif key == "DNS Cache Check":
                    ws.cell(row=row, column=1).value = "DNS Cache Check"
                    ws.cell(row=row, column=1).font = Font(bold=True)
                    row += 1
                    if isinstance(value, list) and value:
                        for dns_entry in value:
                            ws.cell(row=row, column=1).value = dns_entry
                            ws.cell(row=row, column=1).font = default_font
                            row += 1
                    else:
                        ws.cell(row=row, column=1).value = value
                        ws.cell(row=row, column=1).font = default_font
                        row += 1

                # elif key == "USB Malware Protection":
                #     ws.cell(row=row, column=1).value = "USB Malware Protection"
                #     ws.cell(row=row, column=1).font = Font(bold=True)
                #     row += 1
                #     ws.cell(row=row, column=1).value = value
                #     ws.cell(row=row, column=1).font = default_font
                #     row += 1

                elif key == "VPN and Proxy Check":
                    ws.cell(row=row, column=1).value = "VPN and Proxy Check"
                    ws.cell(row=row, column=1).font = Font(bold=True)
                    row += 1
                    ws.cell(row=row, column=1).value = value
                    ws.cell(row=row, column=1).font = default_font
                    row += 1

            wb.save(file_path)
            messagebox.showinfo(
                "Export Success", "Analysis results exported successfully."
            )
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    def add_open_tcp_ports_to_excel(self, ws, data, start_row):
        for idx, port in enumerate(data, start=start_row):
            if isinstance(port, dict):
                ws.cell(row=idx, column=1, value=port.get("Local Address", ""))
                ws.cell(row=idx, column=2, value=port.get("Remote Port", ""))
                ws.cell(row=idx, column=3, value=port.get("Local Port", ""))
                ws.cell(row=idx, column=4, value=port.get("Status", ""))
            else:
                ws.cell(row=idx, column=1, value=str(port))
        return start_row + len(data)

    def add_anonymous_connections_to_excel(self, ws, data, start_row):
        for idx, connection in enumerate(data, start=start_row):
            if isinstance(connection, dict):
                ws.cell(row=idx, column=1, value=connection.get("LocalAddress", ""))
                ws.cell(row=idx, column=2, value=connection.get("RemoteAddress", ""))
                ws.cell(row=idx, column=3, value=connection.get("LocalPort", ""))
                ws.cell(row=idx, column=4, value=connection.get("RemotePort", ""))
            else:
                ws.cell(row=idx, column=1, value=str(connection))
        return start_row + len(data)

    def add_firewall_rules_to_excel(self, ws, data, start_row):
        for idx, rule in enumerate(data, start=start_row):
            if isinstance(rule, dict):
                ws.cell(row=idx, column=1, value=rule.get("Profile", ""))
                ws.cell(row=idx, column=2, value=rule.get("Description", ""))
                ws.cell(row=idx, column=3, value=rule.get("Name", ""))
                ws.cell(row=idx, column=4, value=rule.get("Direction", ""))
                ws.cell(row=idx, column=5, value=rule.get("Action", ""))
            else:
                ws.cell(row=idx, column=1, value=str(rule))
        return start_row + len(data)

    def add_user_account_audit_to_excel(self, ws, data, start_row):
        if isinstance(data, dict):
            ws.cell(row=start_row, column=1, value=data.get("FullName", ""))
            ws.cell(row=start_row, column=2, value=data.get("Name", ""))
            last_login = self.converter_data(str(data.get("LastLogon")))
            ws.cell(row=start_row, column=3, value=last_login)
        else:
            ws.cell(row=start_row, column=1, value=str(data))
        return start_row + 1

    def converter_data(self, data_milliseconds):
        milliseconds = int(data_milliseconds.strip("\/Date()\/"))
        seconds = milliseconds / 1000
        readable_date = datetime.datetime.utcfromtimestamp(seconds)
        formatted_date = readable_date.strftime("%Y-%m-%d %H:%M:%S")
        return formatted_date


if __name__ == "__main__":
    root = tk.Tk()

    windll.shcore.SetProcessDpiAwareness(1)
    app = SystemAnalyzerApp(root)
    root.mainloop()
